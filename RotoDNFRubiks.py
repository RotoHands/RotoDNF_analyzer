from decode_gan import aes128, decData
from bleak import BleakClient
import openpyxl
import keyboard
import websockets
import webbrowser
from  algClass import Alg
import os
import subprocess
import time
import asyncio
from trainer import Trainer, get_new_moves
import msvcrt as m
import pyperclip
import socket
import kociemba

### Setup!!! ###
exeOrder = "EdgesCorners"
import json

from bleak import BleakClient
import bleak
import asyncio
from bleak import BleakScanner

def parse_move(msgLen, value):

    axisPerm = [5, 2, 0, 3, 1, 4]
    facePerm = [0, 1, 2, 5, 8, 7, 6, 3]
    faceOffset = [0, 0, 6, 2, 0, 0]
    curBatteryLevel = -1
    for i in range(0,msgLen,2):

        if(msgLen == 10):
            print(toHexVal(value))
        axis = axisPerm[value[3 + i] >> 1]
        power =[0, 2][value[3 + i] & 1]
        m = axis * 3 + power
        s = ("URFDLB"[axis] + " 2'"[power])
        return (s)

def parseData(value):
    if (len(value) < 4) :
        return None
    if value[0] != 0x2a or value[len(value) - 2] != 0x0d or value[len(value) - 1] != 0x0a:
        return None
    msgType = value[2]
    msgLen = len(value) - 6
    if (msgType == 1):
        return parse_move(msgLen, value)
def toHexVal(value):
    valhex = []
    for i in range(len(value)):
        valhex.append(value[i] >> 4 & 0xf)
        valhex.append(value[i] & 0xf)
    return valhex



async def connect_to_device(address, Cube):
    print("starting", address, "loop")
    notify_uuid = '6e400003-b5a3-f393-e0a9-e50e24dcca9e'
    write_uuid = '6e400002-b5a3-f393-e0a9-e50e24dcca9e'
    service = '6e400001-b5a3-f393-e0a9-e50e24dcca9e'

    async with BleakClient(address, timeout=15.0) as client:
        print("connect to", address)

        try:
            print(await client.get_services())
            await client.write_gatt_char(write_uuid, [51])
            await client.start_notify(notify_uuid,Cube.callback)
            await asyncio.sleep(3000.0)
        except Exception as e:
            print(e)
    print("disconnect from", address)
addr = "ec:6a:31:5b:17:2d"
addrs = "DA:82:22:7A:A8:82"


class DNFanalyzer:
    def __init__(self):
        self.facelet_moves = ""
        self.fail_reason = ""
        self.frame_rate = 0
        self.url = ""
        self.alg = Alg("")
        self.scrambleToExe = ""
        self.scrambleApplied = ""
        self.scramble = True
        self.memo = False
        self.solving = False
        self.startMemoTime = time.time()
        self.memoTime = time.time()
        self.startExeTime = time.time()
        self.exeTime = time.time()
        self.Recon = ""
        self.secMistake = 0.0
        self.success = False
        self.solveTime = 0
        self.exeMoves = []#[move, time, cornersSolved, edgesSolved, index]
        self.scrambleRow = 0
        self.parity = False
        self.solveNum = 0
        self.solution = ""
        self.moveWrong = 0
        self.ws_rec = None
        self.ws_rec_ip = "127.0.0.1"
        self.ws_rec_port = 12345
        self.ws_play = None
        self.ws_play_ip = "127.0.0.1"
        self.ws_play_port = 12321
        self.ws_ui = None
        self.video_time = 0
        self.scramble_moves = []
        self.moves = []
    def resetCube(self):
        self.solving = False
        self.moves = []
        self.moveNumber = 0
        self.alg.reset()
        self.moveNumber = 0
        self.exeMoves = []
        # self.scrambleToExe = getScramble(self)
        self.scrambleApplied = ""
        self.scramble = True
        self.solving = False
        self.secMistake = 0.0
        self.startMemoTime = time.time()
        self.memoTime = time.time()
        self.startExeTime = time.time()
        self.exeTime = time.time()
        self.success = False
        self.rawSol = []
        self.mostPiecesSolved = 0
        self.solution = ""
        self.moveWrong = 0

    def callback(self, sender: int, data: bytearray):
        Cube.moves = parseData(data)

    async def send_ui(self, field, msg):
        data_to_ui = json.dumps({field:msg})
        await self.ws_ui.send(data_to_ui)
    def wait(self):
        while(not keyboard.is_pressed(' ')):
            pass

Cube = DNFanalyzer()
loop = asyncio.get_event_loop()
loop.run_until_complete(connect_to_device(addr, Cube))


def init_ws(ip, port):
    time.sleep(2)
    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    client.connect((ip, port))
    return client



def url(Cube, scramble):#gen url of solve to algcubing
    sol = str(Cube.solution)
    scramble = str(scramble)
    scramble = scramble.replace('\'', "-")
    scramble = scramble.replace(" ", "_")
    newSol = "https://alg.cubing.net/?setup="
    newSol += scramble
    newSol += "%0A&alg="
    solWrong = ""
    allSol = sol.split(" ")
    count = 0
    if (Cube.moveWrong == 0):
        sol = sol.replace("\'", "-")
        sol = sol.replace(" ", "_")
        newSol += sol
    else:
        for x in allSol:

            x = x.replace("\'", "-")
            solWrong += x +"_"
            if(count == Cube.moveWrong):
                solWrong+="%2F%2Fmistake_from_here%0A"
            count+=1

        newSol+=solWrong

    newSol+="%0A"
    return newSol

def playVid(Cube):#play video of so;ve from the sec of mistake

    subprocess.Popen(["python", 'playVid.py'])

    Cube.ws_play = init_ws(Cube.ws_play_ip, Cube.ws_play_port)
    Cube.ws_play.recv(1024).decode('utf-8')
    Cube.ws_play.send(bytearray(str("{}:{}:{}:{}".format("%.2f"%Cube.secMistake, Cube.scrambleRow, Cube.frame_rate, Cube.video_time)), 'utf-8'))


    data = Cube.ws_play.recv(1024).decode('utf-8')
    if data != "finish":
        data_split = data.split(":")
        Cube.secMistake = int(float(data_split[0]))
        Cube.fail_reason = data_split[1]
def getScramble(Cube):# returns Cube object of scramble

    wb = openpyxl.load_workbook(os.getcwd() + "\RotoDNFStats.xlsx", data_only="yes")
    ws = wb["RotoStats"]
    i=1
    while(ws.cell(i,2).value != "no"):
        i+=1
    ws.cell(i,2).value = "yes"
    wb.save(os.getcwd() + "\RotoDNFStats.xlsx")
    Cube.scrambleRow = i
    Cube.scrambleApplied = ws.cell(i,1).value
    return ws.cell(i,1).value

def saveSolve(Cube):#save results to excel
    memo = Cube.memoTime
    exce = Cube.exeTime
    all = memo + exce
    print("all: ", all, ", memo: ", memo, ", exe: ", exce)
    a = input("Solved? (y/n)")#make sure the result is right
    if(a == "y"):
        succsses = "yes"
        Cube.secMistake = 0.0
    else:
        succsses = "no"
    row = Cube.scrambleRow

    wb = openpyxl.load_workbook(os.getcwd() + "\RotoDNFStats.xlsx", data_only="yes")
    ws = wb["RotoStats"]
    strRow = str(Cube.scrambleRow)
    ws.cell(row, 3).value = round(memo,2)
    ws.cell(row, 4).value = round(exce,2)
    ws.cell(row, 5).value = round(all,2)
    ws.cell(row, 6).value = round(Cube.secMistake, 2)
    ws.cell(row, 7).value = succsses

    ws.cell(row,10).value = url(Cube,ws.cell(row,1).value)
    print(Cube.fail_reason)
    ws.cell(row,11).value = Cube.fail_reason
    Cube.url = ws.cell(row,10).value
    pyperclip.copy(Cube.url)
    #ws.cell(row, 8).value = "=HYPERLINK(" + "\"C:\Python\PythonWork\BLD\RotoBLD\RotoDNF\\videos\solve" + strRow +".avi\""+", "+"\"solve"+strRow+"\")"
    ws.cell(row, 9).value =  Cube.solution
    wb.save(os.getcwd() + "\RotoDNFStats.xlsx")
def takeCorners(elem):
    return elem[2]

def takeEdges(elem):
    return elem[3]
def takeIndex(elem):
    return elem[4]
def takeTime(elem):
    return elem[1]


def moves_to_string(moves):
    st = ""
    for m in moves:
        st += m + " "
    return st
async def reset_scramble(Cube, trainer):
    Cube.facelet_moves = ""
    Cube.resetCube()
    Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    await Cube.send_ui("msg", "press space when cube is solved")  # add this to websocket
    Cube.wait()
    # while (not keyboard.is_pressed(' ') == False):
    #     pass
    trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
    trainer.data_move_counter = trainer.data[12]
    await Cube.send_ui("msg", "scramble")
    await Cube.send_ui("scramble", Cube.scrambleToExe)  # add this to websocket
    Cube.scramble_moves = []

def get_facelet_strnig(value):
    state = []
    for i in range(0, len(value) - 2, 3):
        face = value[i ^ 1] << 16 | value[i + 1 ^ 1] << 8 | value[i + 2 ^ 1]
        for j in range(21, -1, -3):
            state.append("URFDLB"[face >> j & 0x7])
            if (j == 12):
                state.append("URFDLB"[int(i / 3)])

    latestFacelet = "".join(state)
    return latestFacelet
async def get_state(Cube, trainer):
    trainer.state_data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f2), trainer.decoder)
    value = trainer.state_data
    print(trainer.last_state_string)
    if value != trainer.last_state_data:
        current_state = get_facelet_strnig(value)
        Cube.facelet_moves += " " + kociemba.solve(trainer.last_state_string, current_state)
        await Cube.send_ui("facelet", Cube.facelet_moves)  # add this to websocket
        trainer.last_state_data = value
        trainer.last_state_string = current_state

async def scramble_cube(Cube, trainer):

    print("hereeee")
    await Cube.send_ui("msg","press space when cube is solved") # add this to websocket
    Cube.wait()
    # while (not keyboard.is_pressed(' ') == False):
    #     pass
    trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
    trainer.data_move_counter = trainer.data[12]
    trainer.state_data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f2), trainer.decoder)
    trainer.last_state_string = get_facelet_strnig(trainer.state_data)

    await Cube.send_ui("msg", "scramble")
    await Cube.send_ui("scramble",Cube.scrambleToExe) # add this to websocket
    Cube.scramble_moves = []
    while (Cube.scramble):  # Scramble Cube
        if (keyboard.is_pressed('q')):
            await reset_scramble(Cube,trainer)
        trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
        trainer.new_moves = get_new_moves(trainer.data, trainer.data_move_counter)
        await get_state(Cube,trainer)
        if (trainer.new_moves):
            newM = trainer.new_moves
            trainer.new_moves.reverse()
            trainer.moves += trainer.new_moves
            newMoves = trainer.new_moves
            for move in newMoves:
                Cube.scramble_moves.append(move)
                await Cube.send_ui("moves",moves_to_string(Cube.scramble_moves))
                # print(move, flush=True, end=" ") # add to websocket
                Cube.alg.movesToExecute = move
                Cube.alg.executeAlg()
                if (Cube.alg.checkSolved() == True):
                    Cube.scramble = False
        trainer.data_move_counter = trainer.data[12]


async def initCube(websocket, path):

    Cube = DNFanalyzer()
    Cube.ws_ui = websocket
    Cube.resetCube()
    Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    trainer = Trainer()
    # trainer.ble_server = BleakClient(trainer.addr)
    # await trainer.ble_server.connect()

    while True:
        Cube.resetCube()
        Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
        Cube.alg.executeAlg()
        Cube.alg.reverseSelf()
        await scramble_cube(Cube,trainer)
        await Cube.send_ui("msg","press enter to start memo") # add to websocket
        Cube.wait()
        Cube.solveTime = time.time()
        subprocess.Popen(["python","recordVid.py"])#record video while solving
        Cube.ws_rec = init_ws(Cube.ws_rec_ip, Cube.ws_rec_port)
        Cube.ws_rec.send(bytearray(str(Cube.scrambleRow), 'utf-8'))
        started = Cube.ws_rec.recv(1024).decode('utf-8')
        diff = time.time() - Cube.solveTime

        print("time difff : {}".format(diff))
        await Cube.send_ui("msg","press enter to start solve") # add to websocket
        Cube.wait()
        time.sleep(1)
        await Cube.send_ui("msg","Solve!")
        Cube.memoTime = time.time() - Cube.solveTime
        moveCount = 0
        Cube.exeTime = time.time() - diff
        exeMoves = []
        moves = []

        while (keyboard.is_pressed(' ') == False):  #While you are solving the cube
            trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
            trainer.new_moves = get_new_moves(trainer.data, trainer.data_move_counter)

            if (trainer.new_moves):
                newMoves = trainer.new_moves

                for move in newMoves:
                    print("move : {} ".format(move))
                    exeMoves.append([move, int(time.time() - Cube.solveTime - diff), 0, 0, moveCount])
                    moveCount += 1
                    moves.append(move)
                    # print(move, flush=True, end=" ")  # add to websocket
                    await Cube.send_ui("moves",moves_to_string(moves))
            trainer.data_move_counter = trainer.data[12]

        await Cube.send_ui("msg","finish")
        Cube.exeTime = time.time()- Cube.exeTime
        Cube.ws_rec.send(bytearray('STOP', 'utf-8'))
        Cube.frame_rate = round(float(Cube.ws_rec.recv(1024).decode('utf-8')))
        Cube.video_time = round(float(Cube.ws_rec.recv(1024).decode('utf-8')))

        #finished solve
        Cube.alg.reset()
        Cube.alg.executeAlgWithMoves(Cube.scrambleToExe)
        Cube.alg.reverseSelf()
        if(Cube.alg.cornerEven() == True):
            Cube.parity = False
            print("no parity")
        else:
            Cube.parity = True
            print("parity")
        #Cube.alg.printState()
        for move in exeMoves:
            Cube.alg.movesToExecute = move[0]
            Cube.alg.reverseSelf()
            Cube.alg.executeAlg()
            Cube.alg.reverseSelf()
            # print("move is :", move[0]," moveNum ", move[4])
            # Cube.alg.printState()#correct  s
            m = move[0]
            t = move[1]
            moveC = move[4]
            corSol = Cube.alg.countSolvedCor()#correct
            edSol = Cube.alg.countSolveEdges()#correct
            Cube.exeMoves.append([m,t,corSol,edSol,moveC])#correct
        sortedByIndex = Cube.exeMoves.copy()
        sortedByIndex.reverse()
        Cube.exeMoves.sort(key=takeEdges, reverse=True)
        sortedByEdges = Cube.exeMoves.copy()  # works
        Cube.exeMoves.sort(key=takeCorners, reverse=True)
        sortedByCorners =  Cube.exeMoves.copy()
        Cube.solution = sortedByIndex.copy()

        print("sortedByEdges ", *sortedByEdges, sep="\n")
        print("sortedByCorners ", *sortedByCorners, sep="\n")
        print("sortedByIndex ", *sortedByIndex, sep="\n")

        solve = ""
        sortedByIndex.reverse()
        for x in sortedByIndex:
            solve += " " + x[0]

        # solved :works
        # edges solved, corners not , no parity : works
        # edges solved corners not, with parity : ???
        # edges not solved, no paritu :??? - - multiple max edges doesnt work
        # edges not solved, parity :???


        Cube.solution = solve
        sortedByIndex.reverse()
        if(exeOrder == "EdgesCorners"):
            if(sortedByIndex[0][2] == 8 and sortedByIndex[0][3] == 12):
                print("Solved!") # works
            elif((sortedByEdges[0][3] == 12 and Cube.parity == False) or (sortedByEdges[0][3] >=10 and Cube.parity == True)):# if there is parity' and max edges >= 10 then it means edges are good


                print("you solved all edges correctly")
                # corners are wrong
                maxCornersFullEdges = []
                for x in sortedByCorners:
                    if((x[3] == 12 and Cube.parity == False) or (x[3] >= 10 and Cube.parity == True)):
                        maxCornersFullEdges.append(x)

                tempSortedIndex = sortedByIndex.copy()
                maxCornersFullEdges.sort(key=takeTime, reverse=True)
                print("maxCornersFullEdges ", *maxCornersFullEdges, sep="\n")
                if(tempSortedIndex[0][3] != 12):#if all edges are not right at the final state, then take the last time all edges were good( mistake in excecution
                    mistakeMove =  maxCornersFullEdges[0]
                else:#take the first time all edges were good (did the wrong Alg)
                    maxCornersFullEdges.reverse()
                    mistakeMove = maxCornersFullEdges[0]
                Cube.moveWrong = mistakeMove[4]
                # Cube.secMistake = round(mistakeMove[1]-Cube.solveTime,2) #the time of the first time you had max corners solved with full edges
                Cube.secMistake = mistakeMove[1]
                print("mistake time was ",Cube.secMistake  , " seconds from the begining")
                print("the last move right  ", mistakeMove[4], "move, which was ", mistakeMove[0][0])
                playVid(Cube)

            else: #not all edges are good find the most edges solved
                max = sortedByEdges[0][3]
                maxEdgesSolved = []
                for move in sortedByEdges:
                    if (move[3] == max):
                        maxEdgesSolved.append(move)
                maxEdgesSolved.sort(key=takeTime)
                sortedByMaxEdgesTime = maxEdgesSolved.copy()
                Cube.moveWrong = sortedByMaxEdgesTime[0][4]
                # Cube.secMistake = round(sortedByMaxEdgesTime[0][1]-Cube.solveTime, 2) # max edges solved first time
                Cube.secMistake = sortedByMaxEdgesTime[0][1]
                print("mistake time was ",Cube.secMistake, " seconds from the begining")
                print("the last move right  ", sortedByMaxEdgesTime[0][4], "move, which was ", sortedByMaxEdgesTime[0][0])
                playVid(Cube)

        else:
            if (sortedByIndex[0][2] == 8 and sortedByIndex[0][3] == 12):
                print("Solved!")

            elif ((sortedByCorners[0][2] == 8 and Cube.parity == False) or (sortedByCorners[0][2] >=6 and Cube.parity == True)):
                print("you solved all corners correctly")
                #not all edges are good

                maxEdgesFullCorners = []#edges are wrong

                for x in sortedByEdges:
                    if ((x[2] == 8 and Cube.parity == False) or (x[2] >= 6 and Cube.parity == True)):
                        maxEdgesFullCorners.append(x)
                maxEdgesFullCorners.sort(key = takeEdges, reverse=True)  # all moves that edges was max
                max = maxEdgesFullCorners[0][2]
                maxEdgesSolved = []

                maxEdgesFullCorners.sort(key=takeTime, reverse=True)
                tempSortedIndex = sortedByIndex.copy()
                if (tempSortedIndex[0][3] != 8):  # if all edges are not right at the final state, then take the last time all edges were good
                    mistakeMove = maxEdgesFullCorners[0]
                else:  # take the first time all edges were good
                    maxEdgesFullCorners.reverse()
                    mistakeMove = maxEdgesFullCorners[0]
                Cube.moveWrong = mistakeMove[4]
                sortedByMaxEdgesTime = maxEdgesFullCorners.copy()
                Cube.secMistake =round(mistakeMove[1]-Cube.solveTime ,2)# the time of the first time you had max corners solved
                print("mistake time was ", Cube.secMistake, " seconds from the begining")
                print("last move right ", mistakeMove[4], "move, which was ",
                      mistakeMove[0])

                playVid(Cube)
            else:  # not all corners are good
                max = sortedByCorners[0][2]
                maxCornersSolved = []
                for move in sortedByEdges:
                    if (move[2] == max):
                        maxCornersSolved.append(move)
                maxCornersSolved.sort(key=takeTime)
                sortedByMaxCornersTime = maxCornersSolved.copy()
                Cube.moveWrong = sortedByMaxCornersTime[0][4]
                Cube.secMistake = sortedByMaxCornersTime[0][1] # the time of the first time you had max corners solved
                print("mistake time was ", Cube.secMistake , " seconds from the begining")
                print("last move right ", sortedByMaxCornersTime[0][4], "move, which was ", sortedByMaxCornersTime[0][4])
                playVid(Cube)
        saveSolve(Cube)


start_server = websockets.serve(initCube, "10.0.0.12", 56789)
#TODO: fix parity, use websocket for ui , use ffmpeg to add metadata of mistakes, multiple, reset ehwn scramble, multiple tries
loop = asyncio.get_event_loop()
loop.run_until_complete(start_server)
loop.run_forever()
