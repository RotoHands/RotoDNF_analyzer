#!/usr/bin/python3.7
from decode_gan import aes128, decData
from bleak import BleakClient
import openpyxl
import keyboard
import websockets
import pickle
from datetime import datetime
from  algClass import Alg
from pathlib import Path
import os
import re
import subprocess
import time
import asyncio
from trainer import Trainer, get_new_moves
import pyperclip
import socket
import kociemba
from dotenv import load_dotenv
from BLD_solve_parse import parse_solve, parse_smart_cube_solve, parse_url
from BLD_solve_parse import Cube
### Setup!!! ###
exeOrder = "EdgesCorners"
import json


class DNFanalyzer:
    def __init__(self):
        self.facelet_moves = ""
        self.fail_reason = ""
        self.frame_rate = 0
        self.url = ""
        self.alg = Alg("")
        self.scrambleToExe = ""
        self.scrambleApplied = ""
        self.scramble_row_original = None
        self.scramble = True
        self.memo = False
        self.solving = False
        self.startMemoTime = time.time()
        self.memoTime = time.time()
        self.startExeTime = time.time()
        self.exeTime = time.time()
        self.Recon = ""
        self.secMistake = 0
        self.secMistake_vid = 0
        self.success = False
        self.solveTime = 0
        self.exeMoves = []#[move, time, cornersSolved, edgesSolved, index]
        self.scrambleRow = 0
        self.parity = False
        self.solveNum = 0
        self.solution = ""
        self.moveWrong = 0
        self.ws_rec = None
        self.ws_rec_ip = "127.0.0.1"
        self.ws_rec_port = 12345
        self.ws_play = None
        self.ws_play_ip = "127.0.0.1"
        self.ws_play_port = 12321
        self.ws_ui = None
        self.video_time = 0
        self.scramble_moves = []
        self.solve_moves = []
        self.start_recording_A = None
        self.start_recording_B = None

    def resetCube(self):
        self.solving = False
        self.moves = []
        self.moveNumber = 0
        self.alg.reset()
        self.moveNumber = 0
        self.exeMoves = []
        self.scrambleToExe, self.scrambleRow = getScramble()
        self.scrambleApplied = ""
        self.scramble = True
        self.solving = False
        self.fail_reason = ""
        self.secMistake = 0
        self.secMistake_vid = 0
        self.startMemoTime = time.time()
        self.memoTime = time.time()
        self.startExeTime = time.time()
        self.exeTime = time.time()
        self.success = False
        self.rawSol = []
        self.mostPiecesSolved = 0
        self.solution = ""
        self.moveWrong = 0
        self.scramble_moves = []
        self.solve_moves = []
        self.solve_stats = []

    async def send_ui(self, field, msg):
        data_to_ui = json.dumps({field:msg})
        await self.ws_ui.send(data_to_ui)

    def wait(self):
        while(not keyboard.is_pressed(' ')):
            pass

def init_ws(ip, port):
    time.sleep(5)
    print("trying ro connect")
    client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    client.connect((ip, port))
    return client

def url(Cube, scramble):#gen url of solve to algcubing
    sol = str(Cube.solution)
    scramble = str(scramble)
    scramble = scramble.replace('\'', "-")
    scramble = scramble.replace(" ", "_")
    newSol = "https://alg.cubing.net/?setup="
    newSol += scramble
    newSol += "%0A&alg="
    solWrong = ""
    allSol = sol.split(" ")
    count = 0
    if (Cube.moveWrong == 0):
        sol = sol.replace("\'", "-")
        sol = sol.replace(" ", "_")
        newSol += sol
    else:
        for x in allSol:

            x = x.replace("\'", "-")
            solWrong += x +"_"
            if(count == Cube.moveWrong):
                solWrong+="%2F%2Fmistake_from_here%0A"
            count+=1

        newSol+=solWrong

    newSol+="%0A"
    return newSol

def playVid(Cube):#play video of so;ve from the sec of mistake

    subprocess.Popen(["python", 'playVid.py'])

    Cube.ws_play = init_ws(Cube.ws_play_ip, Cube.ws_play_port)
    Cube.ws_play.recv(1024).decode('utf-8')
    Cube.ws_play.send(bytearray(str("{}:{}:{}:{}".format("%.2f"%Cube.secMistake_vid, Cube.scrambleRow, Cube.frame_rate, Cube.video_time)), 'utf-8'))


    data = Cube.ws_play.recv(1024).decode('utf-8')
    if data != "finish":
        data_split = data.split(":")
        # Cube.secMistake = int(float(data_split[0]))
        Cube.fail_reason = data_split[1]
def getScramble2(Cube):# returns Cube object of scramble

    excel_path = os.path.join(Path(os.getcwd()).parent.absolute(), "RotoDNFStats.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only="yes")
    ws = wb["RotoStats"]
    i=1
    while(ws.cell(i,2).value != "no"):
        i+=1
    ws.cell(i,2).value = "yes"
    wb.save(excel_path)
    Cube.scrambleRow = i
    Cube.scrambleApplied = ws.cell(i,1).value
    return ws.cell(i,1).value
def getScramble():
    with open ("scrambles.txt", "r") as f:
        data = f.read().split("\n")
        scramble =  data[0]
    del(data[0])
    with open ("scrambles.txt", "w") as f:
        for d in data:
            f.write("{}\n".format(d))
    return (scramble.split(")")[1], scramble.split(")")[0])
def copy_str(string):
    if os.name == 'nt':
        pyperclip.copy(string)
    else:
        subprocess.Popen(["python3", "./test.py", "{}".format(string)])


def save_solve(scramble_row, success, memo, exe, exe_neto, exe_pause, mistake_sec, scrmable, solve_str, fail_reason ,video_path, algs):
    excel_path = os.path.join(os.getcwd(), "RotoDNFStats.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only="yes")
    ws = wb["RotoStats"]
    row = 2
    while True:
        if ws.cell(row,1).value == None:
            break
        else:
            row += 1
    ws.cell(row, 1).value = round(memo + exe, 2)
    ws.cell(row, 2).value = round(memo, 2)
    ws.cell(row, 3).value = round(exe, 2)
    ws.cell(row, 4).value = success
    ws.cell(row, 5).value = algs
    ws.cell(row, 6).value = round(exe_neto/(exe_neto + exe_pause), 2)*100
    ws.cell(row, 8).value = round(mistake_sec, 2)
    ws.cell(row, 9).value = fail_reason
    ws.cell(row, 10).value = scrmable
    ws.cell(row, 11).value = solve_str
    print(video_path)
    ws.cell(row, 12).value = "=HYPERLINK(\"{}\", \"Solve_Link_{}\")".format(video_path, scramble_row)
    ws.cell(row, 7).value = time.ctime()

    copy_str(solve_str)
    wb.save(excel_path)

def saveSolve2(Cube):#save results to excel
    excel_path = os.path.join(os.getcwd(), "RotoDNFStats.xlsx")
    memo = Cube.memoTime
    exce = Cube.exeTime
    all = memo + exce
    a = input("Solved? (y/n)")#make sure the result is right
    if(a == "y"):
        succsses = "yes"
        Cube.secMistake = 0.0
    else:
        succsses = "no"
    row = Cube.scrambleRow

    wb = openpyxl.load_workbook(os.path.join(os.getcwd(), "RotoDNFStats.xlsx"), data_only="yes")
    ws = wb["RotoStats"]
    strRow = str(Cube.scrambleRow)
    ws.cell(row, 3).value = round(memo,2)
    ws.cell(row, 4).value = round(exce,2)
    ws.cell(row, 5).value = round(all,2)
    ws.cell(row, 6).value = round(Cube.secMistake, 2)
    ws.cell(row, 7).value = succsses

    ws.cell(row,10).value = url(Cube,ws.cell(row,1).value)
    print(Cube.fail_reason)
    ws.cell(row,11).value = Cube.fail_reason
    Cube.url = ws.cell(row,10).value
    copy_str(Cube.url)
    ws.cell(row, 9).value =  Cube.solution
    wb.save(excel_path)
def takeCorners(elem):
    return elem[2]

def takeEdges(elem):
    return elem[3]
def takeIndex(elem):
    return elem[4]
def takeTime(elem):
    return elem[1]


def moves_to_string(moves):
    st = ""
    for m in moves:
        st += m + " "
    return st
async def reset_scramble(Cube, trainer):
    Cube.resetCube()
    Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    trainer.facelet_last_string = ""
    trainer.facelet_current_state = ""
    Cube.facelet_moves = ""

    await Cube.send_ui("msg", "press space when cube is solved")  # add this to websocket
    await trainer.set_cube_solved()
    # Cube.wait()
    while (not keyboard.is_pressed(' ')):
         pass
    # trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
    # trainer.data_move_counter = trainer.data[12]
    await Cube.send_ui("msg", "scramble")
    await Cube.send_ui("scramble", Cube.scrambleToExe)  # add this to websocket
    Cube.scramble_moves = []
    await trainer.set_cube_solved()
def get_facelet_strnig(value):
    state = []
    for i in range(0, len(value) - 2, 3):
        face = value[i ^ 1] << 16 | value[i + 1 ^ 1] << 8 | value[i + 2 ^ 1]
        for j in range(21, -1, -3):
            state.append("URFDLB"[face >> j & 0x7])
            if (j == 12):
                state.append("URFDLB"[int(i / 3)])

    latestFacelet = "".join(state)
    return latestFacelet

async def get_state(Cube, trainer):
    trainer.state_data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f2), trainer.decoder)
    value = trainer.state_data
    # print(trainer.last_state_string)
    if value != trainer.last_state_data:
        current_state = get_facelet_strnig(value)
        Cube.facelet_moves += " " + kociemba.solve(trainer.last_state_string, current_state)
        await Cube.send_ui("facelet", Cube.facelet_moves)  # add this to websocket
        trainer.last_state_data = value
        trainer.last_state_string = current_state

async def scramble_cube(Cube, trainer):

    print("next_solve")
    trainer.facelet_last_string = ""
    trainer.facelet_current_state = ""
    Cube.facelet_moves = ""
    await Cube.send_ui("msg","press space when cube is solved")
    await Cube.send_ui("moves", "")
    while (not keyboard.is_pressed(' ')):
         pass
    await trainer.set_cube_solved()
    # trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
    # trainer.data_move_counter = trainer.data[12]
    # trainer.state_data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f2), trainer.decoder)
    # trainer.last_state_string = get_facelet_strnig(trainer.state_data)
    await Cube.send_ui("msg","wait for scramble")
    Cube.scramble_row_original = Cube.scrambleRow
    subprocess.Popen(["python", "./recordVid.py"])  # record video while solving
    await Cube.send_ui("msg","scramble")
    time.sleep(3)

    if os.name == 'nt':
        Cube.ws_rec = init_ws(Cube.ws_rec_ip, Cube.ws_rec_port)
        Cube.ws_rec.send(bytearray(str(Cube.scramble_row_original), 'utf-8'))
        started = Cube.ws_rec.recv(1024).decode('utf-8')
    Cube.start_recfording_A = time.time()
    await Cube.send_ui("msg", "scramble")
    await Cube.send_ui("scramble",Cube.scrambleToExe) # add this to websocket
    Cube.scramble_moves = []

    while (Cube.scramble):  # Scramble Cube
        if (keyboard.is_pressed('q')):
            await reset_scramble(Cube,trainer)
        # trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
        # trainer.new_moves = get_new_moves(trainer.data, trainer.data_move_counter)
        # await get_state(Cube,trainer)
        await asyncio.sleep(0.1)
        # trainer.moves += trainer.new_moves
        if (len(trainer.new_moves) != trainer.data_move_counter):
            trainer.moves += trainer.new_moves
            newMoves = trainer.new_moves[trainer.data_move_counter: len(trainer.new_moves)]
            trainer.data_move_counter = len(trainer.new_moves)
            # trainer.new_moves.reverse()
            for move in newMoves:
                Cube.scramble_moves.append(move)
                await Cube.send_ui("moves",moves_to_string(Cube.scramble_moves))
                Cube.alg.movesToExecute = move
                Cube.alg.executeAlg()
                if (Cube.alg.checkSolved() == True):
                    Cube.scramble = False
            # trainer.new_moves = []
        # trainer.data_move_counter = trainer.data[12]


def parse_to_algs_time(stats, exe_moves):

    mistake_sec = 0
    stats[0]["time"] = exe_moves[0][1]
    for i in range (1,len(stats)):
        stats[i]["time"] = exe_moves[stats[i]["count"] - 1][1]
        if "diff_moves" in stats[i]:
            current = stats[i]["time"]
            first = stats[i - stats[i]["diff_moves"]]["time"]
            time_exe = round(current - stats[1 + i - stats[i]["diff_moves"]]["time"], 2)
            time_alg = round(current - first ,2)
            time_pause = round(time_alg - time_exe, 2)
            alg_str = []
            for j in range(stats[i - stats[i]["diff_moves"]]["count"] + 1, i + 1):
                alg_str.append(stats[j]["move"])
            alg_str = " ".join(alg_str)
            tmp_cube = Cube()
            if stats[i]["piece"] == "edges" or stats[i]["piece"] == "parity":
                alg_str = tmp_cube.parse_alg_to_slice_moves(alg_str)
            stats[i]["alg_stats"] = {"time_alg" : time_alg, "exe" : time_exe, "pause" :  time_pause, "alg_str" :  alg_str}
            stats[i]["alg_stats"]["comment"]  = stats[i]["comment"]
            stats[i]["alg_stats"]["piece"] = stats[i]["piece"]
        mistake_sec_to_end = 0
        if "mistake" in stats[i]["comment"]:
            mistake_sec = exe_moves[stats[i]["count"] - 1][1]
    algs = []
    for move in stats:
        if "diff_moves" in move:
            if "mistake" not in move["comment"]:
                algs.append([move["alg_stats"]])

    return (algs, mistake_sec)

def parse_solve_main(SCRAMBLE, SOLVE, exe_moves,CUBE_SOLVE):

        load_dotenv()  # load .env variable

        cube = parse_solve(SCRAMBLE, SOLVE)

        with open ("failure_pkl.pkl", "wb") as f:
            pickle.dump([cube.solve_stats, exe_moves],f)
        algs_time, mistake_sec = parse_to_algs_time(cube.solve_stats, exe_moves)

        sum_pause = 0
        sum_exe = 0
        for a in algs_time:
            sum_pause += a[0]["pause"]
            sum_exe += a[0]["exe"]
        sum_exe = round(sum_exe, 2)
        sum_pause = round(sum_pause, 2)
        if cube.smart_cube:
            cube = parse_smart_cube_solve(cube)
        cube.time_solve = round(CUBE_SOLVE.memoTime + CUBE_SOLVE.exeTime, 2)
        mistake_sec_to_end = round(cube.time_solve - sum_exe - sum_pause - CUBE_SOLVE.memoTime, 2)
        success = True if cube.solve_stats[-1]['cor'] == 8 and cube.solve_stats[-1]['ed'] == 12 else False
        try:
            flude = round(sum_exe/(sum_exe + sum_pause)*100,2)
        except:
            print("fail flude")
            flude = 0
        cube.name_of_solve = "{}{}({};{};{}%25{}) : {}".format("DNF - " if not success else "", cube.time_solve,round(CUBE_SOLVE.memoTime, 2), round( CUBE_SOLVE.exeTime, 2), flude,";{}".format(mistake_sec) if mistake_sec != 0 else "",time.ctime())
        solve_str = cube.url
        solve_str = re.sub('&title=[^&]*', '&title={}'.format(cube.name_of_solve), solve_str)
        solve_str = re.sub('&time=[^&]*', '&time={}'.format(cube.time_solve), solve_str)
        copy_str(solve_str)
        return (solve_str, cube.solve_stats,algs_time, mistake_sec, sum_exe, sum_pause, mistake_sec_to_end, success)

def solve_description(algs_time):
    edges_algs = 0
    cor_algs = 0
    twist = 0
    flip = 0

    for alg in algs_time:
        if alg[0]['piece'] == "edges":
            if "flip" in alg[0]['comment']:
                flip += 1
            else:
                edges_algs += 2
        if alg[0]['piece'] == "corners":
            if "twist" in alg[0]['comment']:
                twist += 1
            else:
                cor_algs += 2
        if alg[0]['piece'] == "parity":
            cor_algs += 1

    solve_desc = "{}{}/{}{}".format(edges_algs, "'"*flip, cor_algs, "'"*twist)
    return solve_desc

async def initCube(websocket, path):

    Cube = DNFanalyzer()
    Cube.ws_ui = websocket
    Cube.resetCube()
    Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    trainer = Trainer()
    trainer.ble_server = BleakClient(trainer.rubiks1_addr)

    async with BleakClient(trainer.rubiks1_addr, timeout=20.0) as trainer.ble_server:
        if trainer.rubiks == True:
            await trainer.ble_server.start_notify(14, trainer.callback)
        while True:
            Cube.resetCube()
            Cube.alg.movesToExecute = Cube.scrambleToExe  # put scramble moves ready to execute
            Cube.alg.executeAlg()
            Cube.alg.reverseSelf()
            await scramble_cube(Cube,trainer) #also starts recording
            await Cube.send_ui("msg","press space to start memo") # add to websocket
            Cube.wait()
            Cube.start_recording_B = time.time()
            Cube.solveTime = time.time()
            await Cube.send_ui("msg","memo") # add to websocket
            while(len(trainer.new_moves) == trainer.data_move_counter):
                await asyncio.sleep(0.1)
            await Cube.send_ui("msg","Solve!")
            Cube.memoTime = time.time() - Cube.solveTime
            moveCount = 0
            Cube.exeTime = time.time()
            exeMoves = []
            moves = []

            while not keyboard.is_pressed(' '):  #While you are solving the cube
                # trainer.data = decData(await trainer.ble_server.read_gatt_char(trainer.chrct_uuid_f5), trainer.decoder)
                # trainer.new_moves = get_new_moves(trainer.data, trainer.data_move_counter)
                await asyncio.sleep(0.1)
                if (len(trainer.new_moves) != trainer.data_move_counter):
                    trainer.moves += trainer.new_moves
                    newMoves = trainer.new_moves[trainer.data_move_counter : len(trainer.new_moves)]
                    trainer.data_move_counter = len(trainer.new_moves)

                    for move in newMoves:
                        Cube.solve_moves.append(move)
                        exeMoves.append([move, round((time.time() - Cube.solveTime),2), moveCount])
                        moveCount += 1
                        moves.append(move)
                        await Cube.send_ui("moves",moves_to_string(moves))
                # trainer.data_move_counter = trainer.data[12]

            Cube.exeTime = time.time()- Cube.exeTime
            await Cube.send_ui("msg","finish")
            if os.name == "nt":
                Cube.ws_rec.send(bytearray('STOP', 'utf-8'))
                Cube.frame_rate = round(float(Cube.ws_rec.recv(1024).decode('utf-8')))
                Cube.video_time = round(float(Cube.ws_rec.recv(1024).decode('utf-8')))

            scramble = Cube.scrambleToExe
            solve = " ".join(Cube.solve_moves)
            openPath = os.path.join(Path(os.getcwd()).parent.absolute(), "Videos","{}{}".format(str(Cube.scramble_row_original), ".mkv"))
            solve_str, solve_stats, algs_time, mistake_sec, sum_exe, sum_pause, mistake_sec_to_end, success = parse_solve_main(scramble, solve, exeMoves,Cube)
            solve_desc = solve_description(algs_time) if success else None

            Cube.secMistake = mistake_sec
            Cube.secMistake_vid = 0 if mistake_sec == 0 else round(Cube.start_recording_B - Cube.start_recording_A + Cube.secMistake, 2)
            if success == False:
                if os.name != "nt":
                    Cube.fail_reason = input("Enter reason for failure : T-trace, E-exe, M-memo\n")
                    if Cube.fail_reason == "M":
                        Cube.fail_reason = "memo_forgot_error"
                    if Cube.fail_reason == "T":
                        Cube.fail_reason = "trace_error"
                    if Cube.fail_reason == "E":
                        Cube.fail_reason =  "exe_error"
            else:
                Cube.fail_reason = "Success"
            if os.name == "nt":
                playVid(Cube)

            time.sleep(1)
            if os.name == "nt":
                changed_path = os.path.join(Path(os.getcwd()).parent.absolute(), "Videos",   "{}_{}({})_{}{}.mkv".format(Cube.scramble_row_original, round(Cube.memoTime + Cube.exeTime, 2), round(Cube.memoTime), "" if Cube.secMistake == 0 else "{}-({})".format("_mistake", round(Cube.start_recording_B - Cube.start_recording_A + Cube.secMistake, 2)), Cube.fail_reason))
                os.rename(openPath, changed_path)

            with open ("solves.pkl", "rb") as f:
                solves = pickle.load(f)
                solves.append({"date" : datetime.now(), 'success' : success,"solve_time" : Cube.memoTime + Cube.exeTime,"memo_time" :  Cube.memoTime, "exe_time" :  Cube.exeTime, "exe_neto" : sum_exe, "pause_time" : sum_pause,"mistake_sec_to_end" : mistake_sec_to_end,  "fail_reason" : Cube.fail_reason,"algs_time" : algs_time,  "video_path" : changed_path if os.name == "nt" else "" , "scramble" : Cube.scrambleToExe, "solve" :  solve_str, "stats" : solve_stats, "mistake_sec" : mistake_sec})
            with open ("solves.pkl", "wb") as f:
                pickle.dump(solves, f)
            save_solve(Cube.scrambleRow, success, Cube.memoTime,Cube.exeTime, sum_exe, sum_pause, mistake_sec, scramble, solve_str, Cube.fail_reason, changed_path if os.name == "nt" else "", solve_desc )


def main1():

    start_server = websockets.serve(initCube, "127.0.0.1", 5678)
    loop = asyncio.get_event_loop()
    loop.run_until_complete(start_server)
    loop.run_forever()

if __name__ == '__main__':
    main1()
